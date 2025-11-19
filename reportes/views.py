from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta

from proyectos.models import Proyecto, Tarea
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


@login_required
def reportes_index(request):
    """
    Página principal de reportes
    """
    proyectos = Proyecto.objects.all()
    
    context = {
        'proyectos': proyectos,
    }
    
    return render(request, 'reportes/index.html', context)


@login_required
def reporte_proyecto_pdf(request, proyecto_id):
    """
    Genera un reporte PDF completo de un proyecto
    """
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    tareas = proyecto.tareas.all().order_by('-fecha_creacion')
    
    # Crear el objeto BytesIO
    buffer = BytesIO()
    
    # Crear el PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12,
    )
    
    # Título
    title = Paragraph(f"Reporte: {proyecto.nombre}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Información del proyecto
    info_proyecto = [
        ['Campo', 'Valor'],
        ['Creado por', str(proyecto.creado_por)],
        ['Fecha Inicio', proyecto.fecha_inicio.strftime('%d/%m/%Y')],
        ['Fecha Fin', proyecto.fecha_fin.strftime('%d/%m/%Y')],
        ['Progreso', f"{proyecto.calcular_progreso()}%"],
        ['Estado', 'Atrasado' if proyecto.esta_atrasado() else 'En tiempo'],
    ]
    
    tabla_info = Table(info_proyecto, colWidths=[2*inch, 4*inch])
    tabla_info.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(tabla_info)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Descripción
    if proyecto.descripcion:
        desc_heading = Paragraph("Descripción", heading_style)
        elements.append(desc_heading)
        desc = Paragraph(proyecto.descripcion, styles['BodyText'])
        elements.append(desc)
        elements.append(Spacer(1, 0.3 * inch))
    
    # Estadísticas de tareas
    stats_heading = Paragraph("Estadísticas de Tareas", heading_style)
    elements.append(stats_heading)
    
    total_tareas = tareas.count()
    completadas = tareas.filter(estado='completada').count()
    en_progreso = tareas.filter(estado='en_progreso').count()
    pendientes = tareas.filter(estado='pendiente').count()
    
    stats_data = [
        ['Métrica', 'Cantidad'],
        ['Total de Tareas', str(total_tareas)],
        ['Completadas', str(completadas)],
        ['En Progreso', str(en_progreso)],
        ['Pendientes', str(pendientes)],
    ]
    
    tabla_stats = Table(stats_data, colWidths=[3*inch, 2*inch])
    tabla_stats.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(tabla_stats)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Lista de tareas
    tasks_heading = Paragraph("Lista de Tareas", heading_style)
    elements.append(tasks_heading)
    
    if tareas.exists():
        tareas_data = [['Título', 'Asignado', 'Estado', 'Prioridad', 'Vencimiento']]
        
        for tarea in tareas:
            tareas_data.append([
                tarea.titulo[:30],
                str(tarea.asignado_a) if tarea.asignado_a else 'Sin asignar',
                tarea.get_estado_display(),
                tarea.get_prioridad_display(),
                tarea.fecha_vencimiento.strftime('%d/%m/%Y'),
            ])
        
        tabla_tareas = Table(tareas_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
        tabla_tareas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(tabla_tareas)
    else:
        no_tasks = Paragraph("No hay tareas en este proyecto.", styles['BodyText'])
        elements.append(no_tasks)
    
    # Pie de página
    elements.append(Spacer(1, 0.5 * inch))
    footer = Paragraph(
        f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(footer)
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener el PDF del buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_proyecto_{proyecto_id}.pdf"'
    response.write(pdf)
    
    return response


@login_required
def reporte_tareas_excel(request):
    """
    Genera un reporte Excel de todas las tareas con filtros
    """
    # Filtros opcionales
    proyecto_id = request.GET.get('proyecto', None)
    estado = request.GET.get('estado', None)
    fecha_desde = request.GET.get('fecha_desde', None)
    fecha_hasta = request.GET.get('fecha_hasta', None)
    
    # Consulta base
    tareas = Tarea.objects.all()
    
    # Aplicar filtros
    if proyecto_id:
        tareas = tareas.filter(proyecto_id=proyecto_id)
    
    if estado:
        tareas = tareas.filter(estado=estado)
    
    if fecha_desde:
        tareas = tareas.filter(fecha_vencimiento__gte=fecha_desde)
    
    if fecha_hasta:
        tareas = tareas.filter(fecha_vencimiento__lte=fecha_hasta)
    
    tareas = tareas.order_by('-fecha_creacion')
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Tareas"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ['ID', 'Título', 'Proyecto', 'Asignado a', 'Estado', 'Prioridad', 'Fecha Vencimiento', 'Fecha Creación']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Datos
    for tarea in tareas:
        ws.append([
            tarea.id,
            tarea.titulo,
            tarea.proyecto.nombre,
            str(tarea.asignado_a) if tarea.asignado_a else 'Sin asignar',
            tarea.get_estado_display(),
            tarea.get_prioridad_display(),
            tarea.fecha_vencimiento.strftime('%d/%m/%Y'),
            tarea.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
        ])
    
    # Ajustar anchos de columna
    column_widths = [5, 30, 25, 20, 15, 15, 18, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Crear respuesta HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_tareas_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


@login_required
def reporte_general_excel(request):
    """
    Genera un reporte Excel general con múltiples hojas
    """
    wb = openpyxl.Workbook()
    
    # Hoja 1: Proyectos
    ws_proyectos = wb.active
    ws_proyectos.title = "Proyectos"
    
    # Encabezados
    headers_proyectos = ['ID', 'Nombre', 'Creado por', 'Fecha Inicio', 'Fecha Fin', 'Progreso %', '¿Atrasado?']
    ws_proyectos.append(headers_proyectos)
    
    proyectos = Proyecto.objects.all()
    for proyecto in proyectos:
        ws_proyectos.append([
            proyecto.id,
            proyecto.nombre,
            str(proyecto.creado_por),
            proyecto.fecha_inicio.strftime('%d/%m/%Y'),
            proyecto.fecha_fin.strftime('%d/%m/%Y'),
            proyecto.calcular_progreso(),
            'Sí' if proyecto.esta_atrasado() else 'No',
        ])
    
    # Hoja 2: Tareas
    ws_tareas = wb.create_sheet(title="Tareas")
    headers_tareas = ['ID', 'Título', 'Proyecto', 'Asignado', 'Estado', 'Prioridad', 'Vencimiento']
    ws_tareas.append(headers_tareas)
    
    tareas = Tarea.objects.all()
    for tarea in tareas:
        ws_tareas.append([
            tarea.id,
            tarea.titulo,
            tarea.proyecto.nombre,
            str(tarea.asignado_a) if tarea.asignado_a else 'Sin asignar',
            tarea.get_estado_display(),
            tarea.get_prioridad_display(),
            tarea.fecha_vencimiento.strftime('%d/%m/%Y'),
        ])
    
    # Aplicar estilos a ambas hojas
    for ws in [ws_proyectos, ws_tareas]:
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar anchos
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Guardar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_general_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response
